import sys
from copy import copy
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / ".codex_deps"))

from openpyxl import load_workbook
from openpyxl.styles import Alignment


TARGET = Path(r"C:\Users\nhatl\OneDrive\Tài liệu\Downloads\TestCase_RestaurantBookingSystem.xlsx")


def tc(desc, prereq, type_, level, steps, expected):
    return {
        "desc": desc,
        "prereq": prereq,
        "type": type_,
        "level": level,
        "steps": steps,
        "expected": expected,
    }


RESTORE = {
    "CUS_SEARCH": ("CUS_SEARCH", [
        tc("Tìm kiếm nhà hàng theo tên hợp lệ", "Khách hàng đã truy cập trang danh sách nhà hàng", "Function", "Critical", ["Nhập tên nhà hàng hợp lệ, ví dụ HotPot", "Nhấn Tìm kiếm"], "Hệ thống hiển thị các nhà hàng đang hoạt động có tên chứa từ khóa phù hợp"),
        tc("Tìm kiếm nhà hàng theo tên không tồn tại", "Khách hàng đã truy cập trang danh sách nhà hàng", "Function", "Medium", ["Nhập tên nhà hàng không có trong hệ thống", "Nhấn Tìm kiếm"], "Hệ thống hiển thị thông báo không tìm thấy nhà hàng"),
        tc("Tìm kiếm theo tên có dấu tiếng Việt", "Có nhà hàng chứa ký tự tiếng Việt trong tên", "Function", "Medium", ["Nhập tên nhà hàng có dấu", "Nhấn Tìm kiếm"], "Kết quả tìm kiếm vẫn hiển thị đúng nhà hàng phù hợp"),
        tc("Tìm kiếm theo tên viết hoa/viết thường", "Có dữ liệu nhà hàng phù hợp", "Boundary", "Low", ["Nhập từ khóa khác kiểu chữ hoa/thường", "Nhấn Tìm kiếm"], "Hệ thống tìm kiếm không phân biệt hoa thường"),
        tc("Tìm kiếm theo loại ẩm thực Lẩu", "Có nhà hàng thuộc cuisine Lẩu", "Function", "High", ["Chọn loại hình ẩm thực Lẩu", "Nhấn Tìm kiếm"], "Danh sách chỉ hiển thị nhà hàng thuộc loại Lẩu"),
        tc("Tìm kiếm theo loại ẩm thực Nướng", "Có nhà hàng thuộc cuisine Nướng", "Function", "High", ["Chọn loại hình ẩm thực Nướng", "Nhấn Tìm kiếm"], "Danh sách chỉ hiển thị nhà hàng thuộc loại Nướng"),
        tc("Xem chi tiết nhà hàng từ kết quả tìm kiếm", "Có ít nhất một kết quả tìm kiếm", "Function", "Medium", ["Nhấn nút Xem chi tiết", "Quan sát trang được mở"], "Hệ thống mở trang menu/chi tiết đúng nhà hàng đã chọn"),
        tc("Đặt bàn từ kết quả tìm kiếm", "Có ít nhất một kết quả tìm kiếm", "Function", "Medium", ["Nhấn nút Đặt bàn", "Quan sát form đặt bàn"], "Hệ thống chuyển sang trang đặt bàn và truyền đúng mã nhà hàng"),
    ]),
    "CUS_BOOKING": ("CUS_BOOK", [
        tc("Đặt bàn với ghi chú hợp lệ", "Khách hàng đang ở trang đặt bàn", "Function", "Medium", ["Chọn nhà hàng", "Nhập thông tin hợp lệ kèm ghi chú", "Kiểm tra bàn", "Chọn bàn"], "Đơn đặt bàn được tạo, ghi chú không làm ảnh hưởng quá trình đặt bàn"),
        tc("Đặt bàn cho số khách đúng bằng sức chứa bàn", "Có bàn trống với sức chứa phù hợp", "Boundary", "High", ["Nhập số khách bằng Capacity của bàn", "Kiểm tra bàn", "Chọn bàn"], "Bàn được hiển thị và cho phép đặt"),
        tc("Không hiển thị bàn có sức chứa nhỏ hơn số khách", "Có bàn trống nhưng Capacity nhỏ hơn số khách", "Boundary", "High", ["Nhập số khách lớn hơn sức chứa bàn", "Nhấn Kiểm tra bàn"], "Bàn không đủ sức chứa không xuất hiện trong danh sách bàn trống"),
        tc("Đặt bàn sau khi đi từ trang tìm kiếm", "Khách hàng nhấn Đặt bàn ở kết quả tìm kiếm", "Integration", "Medium", ["Mở booking.html?id={RestaurantID}", "Kiểm tra nhà hàng được chọn sẵn"], "Combobox nhà hàng tự chọn đúng nhà hàng từ trang tìm kiếm"),
        tc("Hiển thị giờ hoạt động khi chọn nhà hàng", "Có nhà hàng với Opentime và Closetime", "Function", "Medium", ["Chọn nhà hàng", "Quan sát thông tin giờ hoạt động"], "Trang hiển thị đúng giờ mở cửa và đóng cửa của nhà hàng"),
        tc("Không cho đặt khi chưa tải được giờ nhà hàng", "API thông tin nhà hàng lỗi hoặc chưa phản hồi", "Negative", "Medium", ["Chọn nhà hàng", "Nhấn Kiểm tra bàn khi giờ chưa load"], "Frontend cảnh báo chưa load giờ nhà hàng"),
        tc("Hiển thị popup thanh toán sau khi tạo booking", "Đặt bàn tạo Reservation thành công", "Function", "High", ["Chọn bàn trống", "Quan sát popup thanh toán"], "Popup hiển thị mã đơn, thông tin khách, tiền cọc và mã QR"),
        tc("Đóng popup thanh toán", "Popup thanh toán đang hiển thị", "UI", "Low", ["Nhấn nút đóng popup"], "Popup thanh toán đóng, dữ liệu đặt bàn chưa bị thay đổi"),
        tc("Xác nhận đã chuyển khoản thành công", "Có reservation_id và deposit hợp lệ", "Function", "Critical", ["Nhấn Đã chuyển khoản", "Quan sát kết quả"], "Hệ thống xác nhận thanh toán và hiển thị popup đặt bàn thành công"),
        tc("Quay về trang chủ sau khi đặt bàn thành công", "Popup thành công đang hiển thị", "UI", "Low", ["Nhấn Quay về trang chủ"], "Hệ thống điều hướng về home.html"),
        tc("Đặt bàn với số điện thoại sai định dạng", "Khách hàng đang ở form đặt bàn", "Validation", "Medium", ["Nhập SĐT chứa chữ cái", "Nhập các trường khác hợp lệ", "Kiểm tra bàn/đặt bàn"], "Hệ thống cần cảnh báo SĐT không hợp lệ, không lưu dữ liệu sai"),
        tc("Đặt bàn với ngày trong quá khứ", "Khách hàng đang ở form đặt bàn", "Boundary", "High", ["Chọn ngày nhỏ hơn ngày hiện tại", "Nhập thông tin còn lại", "Kiểm tra bàn"], "Hệ thống cần chặn ngày đặt trong quá khứ"),
    ]),
    "CUS_HIS": ("CUS_HIST", [
        tc("Kiểm tra bộ lọc theo trạng thái đơn hàng cũ", "Khách hàng đã đăng nhập và có đơn ở nhiều trạng thái", "Function", "Medium", ["Truy cập trang Lịch sử đặt bàn", "Lần lượt chọn Đã xong, Đã hủy, Chờ xác nhận"], "Danh sách hiển thị đúng các đơn hàng theo trạng thái đã chọn, không lẫn trạng thái khác"),
        tc("Chặn đánh giá khi chưa hoàn thành bữa ăn", "Đơn đặt bàn đang ở trạng thái Confirmed", "Function", "High", ["Mở lịch sử đặt bàn", "Quan sát đơn chưa hoàn thành"], "Hệ thống không hiển thị nút Đánh giá cho đến khi đơn đủ điều kiện đánh giá"),
        tc("Hiển thị thông báo khi khách chưa có lịch sử", "Khách hàng mới chưa từng đặt bàn", "Boundary", "Low", ["Đăng nhập tài khoản mới", "Mở trang lịch sử"], "Hệ thống hiển thị thông báo không có dữ liệu"),
        tc("Tìm lịch sử với ký tự đặc biệt", "Khách hàng đã đăng nhập", "Boundary", "Low", ["Nhập @#$% vào ô tìm kiếm", "Quan sát danh sách"], "Hệ thống không lỗi và hiển thị dữ liệu rỗng nếu không khớp"),
        tc("Tìm lịch sử bằng khoảng trắng", "Khách hàng đã đăng nhập", "Boundary", "Low", ["Nhập khoảng trắng vào ô tìm kiếm"], "Hệ thống xử lý an toàn và không làm sai danh sách"),
        tc("Mở popup đánh giá từ đơn đã hoàn thành", "Có đơn đủ điều kiện đánh giá", "Function", "Medium", ["Nhấn nút Đánh giá", "Quan sát popup"], "Popup đánh giá hiển thị đúng mã nhà hàng và mã đơn"),
        tc("Đóng popup đánh giá không lưu dữ liệu", "Popup đánh giá đang mở", "UI", "Low", ["Nhập rating/comment", "Nhấn Đóng"], "Popup đóng và không gửi API đánh giá"),
        tc("Hiển thị đánh giá đã gửi", "Đơn đã được đánh giá trước đó", "Function", "Medium", ["Mở lại lịch sử", "Kiểm tra trạng thái đánh giá"], "Hệ thống nhận biết đơn đã đánh giá và không tạo đánh giá trùng"),
    ]),
    "RES_BOOKING": ("RES_BOOK", [
        tc("Xác nhận yêu cầu đặt bàn thành công", "Tài khoản nhân viên đã đăng nhập, có đơn Chờ xử lý", "Function", "High", ["Truy cập trang quản lý yêu cầu đặt bàn", "Chọn một đơn Chờ xử lý", "Nhấn Xác nhận"], "Trạng thái đơn hàng chuyển sang Đã xác nhận"),
        tc("Từ chối yêu cầu đặt bàn thành công", "Tài khoản nhân viên đã đăng nhập, có đơn Chờ xử lý", "Function", "High", ["Chọn một đơn Chờ xử lý", "Nhấn Từ chối"], "Trạng thái đơn hàng chuyển sang Đã từ chối"),
        tc("Hiển thị danh sách yêu cầu đặt bàn mới nhất", "Nhân viên đã đăng nhập", "Function", "Medium", ["Mở trang xác nhận đặt bàn"], "Danh sách booking của nhà hàng hiển thị theo thứ tự phù hợp"),
        tc("Không xác nhận lại đơn đã bị từ chối", "Có đơn trạng thái Rejected", "Negative", "Medium", ["Chọn đơn đã bị từ chối", "Thử nhấn Xác nhận"], "Hệ thống cần không cho thao tác hoặc hiển thị cảnh báo phù hợp"),
        tc("Không từ chối lại đơn đã xác nhận", "Có đơn trạng thái Confirmed", "Negative", "Medium", ["Chọn đơn đã xác nhận", "Thử nhấn Từ chối"], "Hệ thống cần không cho thao tác hoặc hiển thị cảnh báo phù hợp"),
        tc("Kiểm tra thông tin khách trong yêu cầu đặt bàn", "Có booking trong danh sách", "UI", "Low", ["Mở trang quản lý booking", "Quan sát từng dòng"], "Hiển thị đúng tên khách, SĐT, ngày giờ, số người và bàn"),
        tc("Xóa yêu cầu đặt bàn sau khi xác nhận hủy", "Có booking cần xóa", "Function", "Medium", ["Nhấn Xóa booking", "Xác nhận thao tác"], "Booking bị xóa khỏi danh sách"),
        tc("Hủy thao tác xóa booking", "Có booking cần xóa", "UI", "Low", ["Nhấn Xóa booking", "Chọn Cancel ở hộp xác nhận"], "Booking vẫn còn trong danh sách"),
    ]),
    "RES_MENU": ("RES_MENU", [
        tc("Thêm món ăn mới thành công", "Tài khoản nhà hàng đã đăng nhập, đang ở trang Quản lý thực đơn", "Function", "High", ["Bấm nút Thêm món mới", "Nhập đầy đủ tên món, giá, mô tả và ảnh", "Bấm Lưu"], "Hệ thống thông báo thêm món thành công và món mới xuất hiện trong danh sách"),
        tc("Thêm món ăn thiếu hình ảnh", "Tài khoản nhà hàng đã đăng nhập", "Validation", "Medium", ["Mở form thêm món", "Nhập tên và giá nhưng không chọn ảnh", "Bấm Lưu"], "Hệ thống vẫn lưu nếu ảnh không bắt buộc hoặc hiển thị ảnh mặc định"),
        tc("Thêm món ăn thiếu mô tả", "Tài khoản nhà hàng đã đăng nhập", "Validation", "Low", ["Nhập tên, giá và ảnh", "Bỏ trống mô tả", "Bấm Lưu"], "Món ăn được lưu nếu mô tả không bắt buộc"),
        tc("Sửa mô tả món ăn", "Có món ăn trong thực đơn", "Function", "Medium", ["Chọn món cần sửa", "Cập nhật mô tả", "Bấm Lưu"], "Mô tả món ăn được cập nhật trên danh sách"),
        tc("Sửa ảnh món ăn", "Có món ăn trong thực đơn", "Function", "Medium", ["Chọn món cần sửa", "Chọn ảnh mới", "Bấm Lưu"], "Ảnh món ăn được cập nhật thành ảnh mới"),
        tc("Hủy thao tác sửa món", "Đang mở form sửa món", "UI", "Low", ["Thay đổi dữ liệu trên form", "Nhấn Hủy"], "Thông tin món ăn giữ nguyên như trước khi sửa"),
        tc("Xóa món ăn sau khi xác nhận", "Có món ăn trong thực đơn", "Function", "High", ["Nhấn Xóa món", "Xác nhận xóa"], "Món ăn bị xóa khỏi danh sách thực đơn"),
        tc("Hủy thao tác xóa món", "Có món ăn trong thực đơn", "UI", "Low", ["Nhấn Xóa món", "Chọn Cancel"], "Món ăn vẫn còn trong danh sách"),
        tc("Ẩn món ăn tạm thời", "Có món đang hiển thị", "Function", "Medium", ["Nhấn Ẩn món"], "Món chuyển sang trạng thái ẩn và không hiển thị cho order"),
        tc("Hiện lại món ăn đã ẩn", "Có món đang bị ẩn", "Function", "Medium", ["Nhấn Hiện món"], "Món chuyển sang trạng thái hiển thị"),
        tc("Kiểm tra giá món hiển thị đúng định dạng", "Có món ăn có giá", "UI", "Low", ["Mở trang quản lý thực đơn", "Quan sát giá món"], "Giá món hiển thị đúng định dạng tiền Việt"),
        tc("Không cho truy cập quản lý menu khi chưa đăng nhập", "Không có token đăng nhập", "Security", "High", ["Mở trang menu_mgmt.html"], "Hệ thống yêu cầu đăng nhập hoặc chặn truy cập"),
    ]),
}


def existing_descriptions(ws):
    values = set()
    for row in range(6, ws.max_row + 1):
        val = ws.cell(row, 2).value
        if isinstance(val, str) and val.strip():
            values.add(normalize(val))
    return values


def normalize(text):
    return " ".join(str(text).lower().split())


def next_id(prefix, ws):
    max_no = 0
    marker = prefix + "_"
    for row in range(6, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if isinstance(val, str) and val.startswith(marker):
            tail = val.rsplit("_", 1)[-1]
            if tail.isdigit():
                max_no = max(max_no, int(tail))
    return max_no + 1


def copy_row_style(ws, src_row, dst_row):
    for col in range(1, 12):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = Alignment(vertical="top", wrap_text=True)


def append_case(ws, prefix, case, idx):
    start = ws.max_row + 1
    # Keep one real blank row between current block and restored block if not already blank.
    if any(ws.cell(ws.max_row, c).value for c in range(1, 12)):
        start += 1
    src_row = 6 if ws.max_row >= 6 else 5
    row = start
    for step_no, step in enumerate(case["steps"], start=1):
        copy_row_style(ws, src_row, row)
        if step_no == 1:
            ws.cell(row, 1).value = f"{prefix}_{idx}"
            ws.cell(row, 2).value = case["desc"]
            ws.cell(row, 3).value = case["prereq"]
            ws.cell(row, 4).value = case["type"]
            ws.cell(row, 5).value = case["level"]
            ws.cell(row, 8).value = case["expected"]
        ws.cell(row, 6).value = step_no
        ws.cell(row, 7).value = step
        row += 1

    end = row - 1
    if end > start:
        for col in [1, 2, 3, 4, 5, 8, 9, 10, 11]:
            ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)

    for r in range(start, end + 1):
        ws.row_dimensions[r].height = 28


def main():
    wb = load_workbook(TARGET)
    touched = {}
    for sheet, (prefix, cases) in RESTORE.items():
        ws = wb[sheet]
        seen = existing_descriptions(ws)
        idx = next_id(prefix, ws)
        added = 0
        for case in cases:
            if normalize(case["desc"]) in seen:
                continue
            append_case(ws, prefix, case, idx)
            seen.add(normalize(case["desc"]))
            idx += 1
            added += 1
        ws.auto_filter.ref = f"A5:K{ws.max_row}"
        touched[sheet] = added
    wb.save(TARGET)
    print(touched)


if __name__ == "__main__":
    main()
