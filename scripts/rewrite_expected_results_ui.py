import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / ".codex_deps"))

from openpyxl import load_workbook


TARGET = Path(r"C:\Users\nhatl\OneDrive\Tài liệu\Downloads\TestCase_RestaurantBookingSystem.xlsx")


REPLACEMENTS = {
    "API /login trả 200, lưu access_token, role CUSTOMER và chuyển đến màn khách hàng":
        "Đăng nhập thành công, hệ thống chuyển đến màn hình dành cho khách hàng",
    "Đăng nhập thành công, token chứa role STAFF và restaurant_id":
        "Đăng nhập thành công, hệ thống chuyển đến màn hình dành cho nhân viên nhà hàng",
    "API trả 401 và hiển thị thông báo sai tài khoản hoặc mật khẩu":
        "Hệ thống hiển thị thông báo sai tài khoản hoặc mật khẩu",
    "API trả 401, không sinh token":
        "Hệ thống hiển thị thông báo sai tài khoản hoặc mật khẩu và không cho đăng nhập",
    "Token/role bị xóa, người dùng được đưa về trang đăng nhập":
        "Đăng xuất thành công, người dùng được đưa về trang đăng nhập",
    "API /registerRequest trả 201, tạo user và trả access_token":
        "Đăng ký thành công, hệ thống tạo tài khoản mới và cho phép người dùng tiếp tục sử dụng",
    "API trả 400 với thông báo tên đăng nhập đã tồn tại":
        "Hệ thống hiển thị thông báo tên đăng nhập đã tồn tại và không cho tạo tài khoản",
    "Form/API từ chối dữ liệu hoặc không lưu email sai định dạng":
        "Hệ thống hiển thị thông báo email không hợp lệ và không cho lưu dữ liệu",
    "Hệ thống xử lý an toàn, RestaurantID không hợp lệ không làm lỗi server":
        "Hệ thống hiển thị thông báo thông tin nhà hàng không hợp lệ và không cho lưu dữ liệu",
    "API trả 401, frontend yêu cầu đăng nhập lại":
        "Hệ thống hiển thị thông báo phiên đăng nhập không hợp lệ và yêu cầu đăng nhập lại",
    "API trả 403 và không trả danh sách user":
        "Hệ thống hiển thị thông báo không có quyền truy cập và không hiển thị danh sách người dùng",

    "Frontend/API báo tên là bắt buộc":
        "Hệ thống hiển thị thông báo họ tên là bắt buộc và không cho đặt bàn",
    "API /book trả 400 với thông báo tên và SĐT bắt buộc":
        "Hệ thống hiển thị thông báo tên và số điện thoại là bắt buộc, không cho đặt bàn",
    "API trả lỗi bàn đã được đặt trong khung giờ này":
        "Hệ thống hiển thị thông báo bàn đã được đặt trong khung giờ này",
    "API trả 400 bàn không hợp lệ":
        "Hệ thống hiển thị thông báo bàn không hợp lệ và không cho tạo đặt bàn",
    "Booking lưu UserID của token và xuất hiện trong lịch sử khách hàng":
        "Đơn đặt bàn được lưu cho tài khoản đang đăng nhập và hiển thị trong lịch sử đặt bàn",

    "API /search lọc đúng Restaurant.CuisineID":
        "Hệ thống hiển thị đúng danh sách nhà hàng theo loại ẩm thực đã chọn",
    "API xử lý an toàn, không lỗi 500":
        "Hệ thống xử lý an toàn, không bị treo và hiển thị kết quả phù hợp",
    "API trả dữ liệu rỗng hoặc thông báo không tìm thấy, frontend không crash":
        "Hệ thống hiển thị thông báo không tìm thấy nhà hàng và giao diện vẫn hoạt động bình thường",
    "Frontend hiển thị cảnh báo lỗi gọi API":
        "Hệ thống hiển thị thông báo không thể tải dữ liệu, người dùng không bị mất thao tác hiện tại",

    "API /review tạo Review và trả thông báo thành công":
        "Hệ thống lưu đánh giá và hiển thị thông báo gửi đánh giá thành công",
    "API trả reviewed=true, rating và comment":
        "Hệ thống hiển thị trạng thái đã đánh giá cùng số sao và nhận xét đã gửi",
    "API trả 401 chưa đăng nhập":
        "Hệ thống hiển thị thông báo cần đăng nhập trước khi thực hiện thao tác",
    "Popup đóng và không gửi API đánh giá":
        "Popup đóng lại, đánh giá chưa được lưu và dữ liệu không thay đổi",

    "Booking bị xóa khỏi DB và danh sách cập nhật":
        "Đơn đặt bàn được xóa khỏi danh sách và màn hình cập nhật lại dữ liệu",
    "API trả lỗi Booking not found":
        "Hệ thống hiển thị thông báo không tìm thấy đơn đặt bàn",
    "API trả lỗi Not found":
        "Hệ thống hiển thị thông báo không tìm thấy dữ liệu cần thao tác",
    "API trả Wrong datetime format":
        "Hệ thống hiển thị thông báo ngày giờ không hợp lệ và không cho lưu đặt bàn",
    "API trả 401":
        "Hệ thống yêu cầu đăng nhập trước khi tiếp tục",
    "API trả Table not found":
        "Hệ thống hiển thị thông báo không tìm thấy bàn",

    "API trả Food not found":
        "Hệ thống hiển thị thông báo không tìm thấy món ăn",
    "API trả 404 Not found":
        "Hệ thống hiển thị thông báo không tìm thấy dữ liệu cần thao tác",
    "API trả 403 chỉ nhân viên mới được xem":
        "Hệ thống hiển thị thông báo chỉ nhân viên nhà hàng mới được xem chức năng này",
    "API trả nhân viên chưa được phân công nhà hàng":
        "Hệ thống hiển thị thông báo nhân viên chưa được phân công nhà hàng và không cho thêm món",
    "API không trả món đã ẩn":
        "Món đã ẩn không còn hiển thị trong menu gọi món",

    "API trả Table already exists":
        "Hệ thống hiển thị thông báo số bàn đã tồn tại và không cho lưu bàn mới",
    "Hệ thống không tạo bàn và không lỗi 500":
        "Hệ thống hiển thị thông báo sức chứa không hợp lệ và không tạo bàn mới",

    "API /orders tạo order active và order_item tương ứng":
        "Hệ thống tạo đơn gọi món mới và hiển thị danh sách món đã chọn cho bàn",
    "API trả 404 Table not found":
        "Hệ thống hiển thị thông báo không tìm thấy bàn và không cho gửi order",

    "API không làm lỗi server; cần báo bàn không tồn tại hoặc xử lý rõ ràng":
        "Hệ thống hiển thị thông báo không tìm thấy bàn hoặc hướng dẫn người dùng kiểm tra lại",
    "localStorage currentOrder bị xóa":
        "Danh sách món tạm thời được làm mới sau khi thanh toán thành công",
    "Frontend không báo thành công giả khi API thất bại":
        "Hệ thống hiển thị thông báo thanh toán thất bại và không chuyển trạng thái bàn",
    "API trả lỗi Sai số tiền và không tạo Payment":
        "Hệ thống hiển thị thông báo số tiền không đúng và không xác nhận thanh toán",
    "API trả Not found":
        "Hệ thống hiển thị thông báo không tìm thấy đơn cần thanh toán",
    "API trả Booking already processed":
        "Hệ thống hiển thị thông báo đơn đã được xử lý trước đó và không cho thanh toán lại",
    "API không tạo Payment và trả lỗi dữ liệu":
        "Hệ thống hiển thị thông báo thiếu thông tin thanh toán và không xác nhận thanh toán",
    "API trả Not found hoặc lỗi dữ liệu phù hợp":
        "Hệ thống hiển thị thông báo thiếu mã đặt bàn và không xác nhận thanh toán",
    "Tiền cọc hiển thị bằng deposit backend trả về":
        "Tiền cọc hiển thị đúng với số tiền hệ thống đã tính cho đơn đặt bàn",

    "API trả 403":
        "Hệ thống hiển thị thông báo không có quyền thực hiện thao tác",
    "API trả cập nhật thành công và danh sách reload":
        "Hệ thống hiển thị thông báo cập nhật thành công và làm mới danh sách",
    "Role trong DB đổi thành STAFF":
        "Vai trò người dùng được cập nhật thành STAFF trên danh sách quản lý",
    "API trả 404 không tìm thấy người dùng":
        "Hệ thống hiển thị thông báo không tìm thấy người dùng",
    "API trả lỗi khi xóa người dùng":
        "Hệ thống hiển thị thông báo không thể xóa người dùng",
    "Response không chứa Password":
        "Danh sách người dùng không hiển thị mật khẩu",
    "Token bị xóa và chuyển về login.html":
        "Đăng xuất thành công, hệ thống chuyển về trang đăng nhập",

    "API trả 201 và cuisine xuất hiện trong danh sách":
        "Hệ thống hiển thị thông báo thêm thành công và cuisine xuất hiện trong danh sách",
    "API trả 400 tên không được để trống":
        "Hệ thống hiển thị thông báo tên không được để trống và không cho lưu",
    "API trả 400 danh mục đã tồn tại":
        "Hệ thống hiển thị thông báo danh mục đã tồn tại và không cho lưu trùng",
    "Hệ thống cần chặn quá độ dài hoặc trả lỗi DB rõ ràng":
        "Hệ thống hiển thị thông báo tên quá dài và không cho lưu dữ liệu",
    "API trả cập nhật thành công":
        "Hệ thống hiển thị thông báo cập nhật thành công",
    "API trả 404 không tìm thấy":
        "Hệ thống hiển thị thông báo không tìm thấy dữ liệu cần thao tác",
    "API yêu cầu JWT hợp lệ":
        "Hệ thống yêu cầu đăng nhập bằng tài khoản có quyền phù hợp",
    "Không gọi API và dữ liệu không đổi":
        "Dữ liệu không thay đổi khi người dùng hủy thao tác",

    "API trả 400 tên nhà hàng không được để trống":
        "Hệ thống hiển thị thông báo tên nhà hàng không được để trống và không cho lưu",
    "API trả 404 không tìm thấy nhà hàng":
        "Hệ thống hiển thị thông báo không tìm thấy nhà hàng",
    "API trả lỗi phù hợp":
        "Hệ thống hiển thị thông báo không thể thực hiện thao tác",
    "Select hiển thị CuisineID/CuisineName từ API":
        "Danh sách loại ẩm thực hiển thị đầy đủ để người dùng chọn",

    "API trả status success, total_report, total_6_months và danh sách nhà hàng":
        "Hệ thống hiển thị tổng doanh thu, doanh thu 6 tháng và danh sách nhà hàng theo bộ lọc",
    "Frontend/API báo thiếu tháng cần xem báo cáo":
        "Hệ thống hiển thị thông báo cần chọn tháng báo cáo",
    "Kết quả rỗng, total = 0, không lỗi server":
        "Hệ thống hiển thị bảng rỗng và tổng doanh thu bằng 0, giao diện vẫn hoạt động bình thường",
    "Hệ thống cần trả lỗi dữ liệu thay vì lỗi 500":
        "Hệ thống hiển thị thông báo tháng báo cáo không hợp lệ",
    "Frontend hiển thị token không hợp lệ hoặc yêu cầu đăng nhập lại":
        "Hệ thống hiển thị thông báo phiên đăng nhập hết hạn và yêu cầu đăng nhập lại",
}


def rewrite_remaining(text):
    text = text.replace("API ", "Hệ thống ")
    text = text.replace("API/", "Hệ thống ")
    text = text.replace("Backend", "Hệ thống")
    text = text.replace("backend", "hệ thống")
    text = text.replace("Frontend", "Giao diện")
    text = text.replace("frontend", "giao diện")
    text = text.replace("Response", "Kết quả")
    text = text.replace("response", "kết quả")
    text = text.replace("token", "phiên đăng nhập")
    text = text.replace("Token", "Phiên đăng nhập")
    text = text.replace("JWT", "phiên đăng nhập")
    text = text.replace("localStorage", "dữ liệu tạm trên trình duyệt")
    text = text.replace("DB", "dữ liệu hệ thống")
    text = text.replace("database", "dữ liệu hệ thống")
    for code in ["200", "201", "400", "401", "403", "404", "500"]:
        text = text.replace(f"trả {code}", "hiển thị thông báo phù hợp")
        text = text.replace(code, "")
    return " ".join(text.split())


def main():
    wb = load_workbook(TARGET)
    changed = {}
    for ws in wb.worksheets:
        if ws.max_row < 5 or ws.cell(5, 8).value != "Expected Results":
            continue
        count = 0
        for row in range(6, ws.max_row + 1):
            cell = ws.cell(row, 8)
            if not isinstance(cell.value, str) or not cell.value.strip():
                continue
            old = cell.value
            new = REPLACEMENTS.get(old, old)
            if new == old:
                cleaned = rewrite_remaining(old)
                banned = ["API", "Backend", "backend", "Response", "response", "token", "Token", "JWT", "401", "403", "404", "500", "localStorage"]
                if any(term in old for term in banned):
                    new = cleaned
            if new != old:
                cell.value = new
                count += 1
        if count:
            changed[ws.title] = count
    wb.save(TARGET)
    print(changed)


if __name__ == "__main__":
    main()
