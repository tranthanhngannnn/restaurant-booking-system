import re
import sys
from copy import copy
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / ".codex_deps"))

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


TARGET = Path(r"C:\Users\nhatl\OneDrive\Tài liệu\Downloads\TestCase_RestaurantBookingSystem.xlsx")

HEADERS = [
    "Test Case ID",
    "Test case Description",
    "PreRequisites",
    "Type",
    "Level",
    "Steps",
    "Execution Steps",
    "Expected Results",
    "Actual Result",
    "Result",
    "Note",
]


def tc(desc, prereq, type_, level, steps, expected):
    return {
        "desc": desc,
        "prereq": prereq,
        "type": type_,
        "level": level,
        "steps": steps,
        "expected": expected,
    }


def common_invalid_login_steps(username, password):
    return [
        "Truy cập trang đăng nhập",
        f"Nhập username: {username}",
        f"Nhập password: {password}",
        "Nhấn nút Đăng nhập",
    ]


SHEET_CASES = {
    "LOGIN_REGISTER": ("USER_AUTH", [
        tc("Đăng nhập thành công với tài khoản CUSTOMER hợp lệ", "Tài khoản CUSTOMER đã tồn tại trong hệ thống", "Function", "Critical", ["Truy cập trang login.html", "Nhập username và password hợp lệ", "Nhấn Đăng nhập"], "API /login trả 200, lưu access_token, role CUSTOMER và chuyển đến màn khách hàng"),
        tc("Đăng nhập thành công với tài khoản STAFF hợp lệ", "Tài khoản STAFF đã được gán RestaurantID", "Function", "Critical", ["Truy cập trang login.html", "Nhập tài khoản STAFF", "Nhấn Đăng nhập"], "Đăng nhập thành công, token chứa role STAFF và restaurant_id"),
        tc("Đăng nhập thành công với tài khoản ADMIN hợp lệ", "Tài khoản ADMIN đã tồn tại", "Function", "Critical", ["Truy cập trang login.html", "Nhập tài khoản ADMIN", "Nhấn Đăng nhập"], "Đăng nhập thành công và truy cập được các màn quản trị"),
        tc("Đăng nhập thất bại khi sai mật khẩu", "Username tồn tại trong bảng users", "Negative", "High", common_invalid_login_steps("customer1", "sai_mat_khau"), "API trả 401 và hiển thị thông báo sai tài khoản hoặc mật khẩu"),
        tc("Đăng nhập thất bại khi username không tồn tại", "Không có user tên unknown_user", "Negative", "High", common_invalid_login_steps("unknown_user", "123456"), "API trả 401, không sinh token"),
        tc("Đăng nhập khi bỏ trống username", "Người dùng đang ở trang login", "Validation", "High", common_invalid_login_steps("(trống)", "123456"), "Không đăng nhập, hiển thị thông báo lỗi phù hợp"),
        tc("Đăng nhập khi bỏ trống password", "Người dùng đang ở trang login", "Validation", "High", common_invalid_login_steps("customer1", "(trống)"), "Không đăng nhập, hiển thị thông báo lỗi phù hợp"),
        tc("Đăng nhập với username có ký tự đặc biệt", "Người dùng đang ở trang login", "Boundary", "Medium", common_invalid_login_steps("admin' OR '1'='1", "123456"), "Hệ thống không lỗi, không bỏ qua xác thực và trả thông báo đăng nhập thất bại"),
        tc("Đăng xuất khỏi tài khoản đã đăng nhập", "Người dùng đã đăng nhập và localStorage có token", "Function", "High", ["Nhấn nút Đăng xuất", "Xác nhận đăng xuất", "Quan sát localStorage và điều hướng"], "Token/role bị xóa, người dùng được đưa về trang đăng nhập"),
        tc("Truy cập màn cần quyền sau khi đăng xuất", "Người dùng vừa đăng xuất", "Security", "High", ["Mở trực tiếp trang admin hoặc history", "Gọi API cần token"], "Bị chặn truy cập hoặc yêu cầu đăng nhập lại"),
        tc("Đăng ký CUSTOMER thành công", "Username, email, phone chưa tồn tại", "Function", "Critical", ["Truy cập register.html", "Nhập username, password, email, phone", "Chọn role CUSTOMER", "Nhấn Đăng ký"], "API /registerRequest trả 201, tạo user và trả access_token"),
        tc("Đăng ký STAFF thành công có RestaurantID hợp lệ", "RestaurantID đã tồn tại hoặc được admin gán", "Function", "High", ["Mở form đăng ký", "Nhập thông tin STAFF và RestaurantID", "Nhấn Đăng ký"], "User STAFF được tạo và lưu RestaurantID dạng số"),
        tc("Đăng ký thất bại khi username đã tồn tại", "Username test đã tồn tại trong bảng users", "Negative", "High", ["Mở trang đăng ký", "Nhập username trùng", "Nhập các trường còn lại hợp lệ", "Nhấn Đăng ký"], "API trả 400 với thông báo tên đăng nhập đã tồn tại"),
        tc("Đăng ký khi bỏ trống username", "Người dùng đang ở trang đăng ký", "Validation", "High", ["Bỏ trống username", "Nhập password, email, phone", "Nhấn Đăng ký"], "Không tạo tài khoản hoặc hiển thị lỗi bắt buộc nhập username"),
        tc("Đăng ký khi bỏ trống password", "Người dùng đang ở trang đăng ký", "Validation", "High", ["Nhập username mới", "Bỏ trống password", "Nhấn Đăng ký"], "Không tạo tài khoản hoặc hiển thị lỗi bắt buộc nhập password"),
        tc("Đăng ký với email sai định dạng", "Người dùng đang ở trang đăng ký", "Validation", "Medium", ["Nhập email dạng abc@", "Nhập các trường khác hợp lệ", "Nhấn Đăng ký"], "Form/API từ chối dữ liệu hoặc không lưu email sai định dạng"),
        tc("Đăng ký với số điện thoại vượt quá 11 ký tự", "Người dùng đang ở trang đăng ký", "Boundary", "Medium", ["Nhập phone 12 chữ số", "Nhập các trường khác hợp lệ", "Nhấn Đăng ký"], "Hệ thống cảnh báo dữ liệu không hợp lệ hoặc không lưu quá độ dài Phone"),
        tc("Đăng ký STAFF với RestaurantID không phải số", "Người dùng đang ở trang đăng ký", "Boundary", "Medium", ["Nhập RestaurantID = ABC", "Nhập thông tin STAFF hợp lệ", "Nhấn Đăng ký"], "Hệ thống xử lý an toàn, RestaurantID không hợp lệ không làm lỗi server"),
        tc("Token đăng nhập hết hạn khi gọi API bảo vệ", "Có token đã hết hạn hoặc bị sửa", "Security", "High", ["Gọi API /api/v1/customer/history bằng token không hợp lệ", "Quan sát phản hồi"], "API trả 401, frontend yêu cầu đăng nhập lại"),
        tc("Role không đúng bị chặn khi truy cập chức năng quản trị", "Đăng nhập bằng CUSTOMER", "Security", "High", ["Dùng token CUSTOMER gọi /api/v1/admin/users", "Quan sát phản hồi"], "API trả 403 và không trả danh sách user"),
    ]),
    "CUS_SEARCH": ("CUS_SEARCH", [
        tc("Tìm kiếm nhà hàng khi không nhập bộ lọc", "Có dữ liệu nhà hàng trong hệ thống", "Function", "High", ["Mở trang search.html", "Để trống ô địa chỉ và cuisine", "Nhấn Tìm kiếm"], "Hiển thị danh sách nhà hàng và trạng thái còn bàn/hết bàn"),
        tc("Tìm kiếm theo địa chỉ hợp lệ", "Có nhà hàng ở địa chỉ cần tìm", "Function", "High", ["Nhập một phần địa chỉ", "Nhấn Tìm kiếm"], "Chỉ hiển thị nhà hàng có Address chứa từ khóa"),
        tc("Tìm kiếm theo cuisine hợp lệ", "Có dữ liệu CuisineID tương ứng", "Function", "High", ["Chọn một loại ẩm thực", "Nhấn Tìm kiếm"], "API /search lọc đúng Restaurant.CuisineID"),
        tc("Tìm kiếm kết hợp địa chỉ và cuisine", "Có nhà hàng thỏa cả hai điều kiện", "Function", "High", ["Nhập địa chỉ", "Chọn cuisine", "Nhấn Tìm kiếm"], "Danh sách kết quả thỏa đồng thời địa chỉ và cuisine"),
        tc("Tìm kiếm không có kết quả", "Không có nhà hàng khớp từ khóa", "Negative", "Medium", ["Nhập từ khóa không tồn tại", "Nhấn Tìm kiếm"], "Hiển thị thông báo không tìm thấy nhà hàng"),
        tc("Tìm kiếm với ký tự đặc biệt", "Người dùng ở trang tìm kiếm", "Boundary", "Medium", ["Nhập @#$% vào ô tìm kiếm", "Nhấn Tìm kiếm"], "Hệ thống không crash và trả kết quả rỗng hoặc thông báo phù hợp"),
        tc("Tìm kiếm với chuỗi rất dài", "Người dùng ở trang tìm kiếm", "Boundary", "Medium", ["Nhập chuỗi trên 200 ký tự", "Nhấn Tìm kiếm"], "API xử lý an toàn, không lỗi 500"),
        tc("Cuisine truyền lên không phải số", "Có thể sửa query string trên trình duyệt", "Negative", "High", ["Gọi /search?cuisine=abc", "Quan sát phản hồi"], "Hệ thống xử lý lỗi dữ liệu lọc, không làm sập ứng dụng"),
        tc("Hiển thị trạng thái còn bàn", "Nhà hàng có ít nhất một bàn Status = Trống", "Function", "Medium", ["Tìm nhà hàng đó", "Quan sát card kết quả"], "Card hiển thị Còn bàn"),
        tc("Hiển thị trạng thái hết bàn", "Tất cả bàn của nhà hàng không ở trạng thái Trống", "Function", "Medium", ["Tìm nhà hàng đó", "Quan sát card kết quả"], "Card hiển thị Hết bàn"),
        tc("Điều hướng sang xem chi tiết/menu", "Kết quả tìm kiếm có ít nhất một nhà hàng", "Function", "Medium", ["Nhấn Xem chi tiết trên một card"], "Chuyển tới menu.html?id={RestaurantID} đúng nhà hàng"),
        tc("Điều hướng sang đặt bàn", "Kết quả tìm kiếm có ít nhất một nhà hàng", "Function", "Medium", ["Nhấn Đặt bàn trên một card"], "Chuyển tới booking.html?id={RestaurantID} và tự chọn nhà hàng tương ứng"),
        tc("Tải trang tìm kiếm khi chưa đăng nhập", "Không có token trong localStorage", "Function", "Low", ["Mở search.html", "Quan sát menu sidebar"], "Trang vẫn cho tìm kiếm, link lịch sử không hiển thị"),
        tc("Tải trang tìm kiếm khi đã đăng nhập", "Token khách hàng hợp lệ", "Function", "Low", ["Mở search.html", "Gọi /customer/me", "Quan sát sidebar"], "Link lịch sử được hiển thị"),
        tc("API restaurants trả danh sách phục vụ combobox", "Có dữ liệu restaurant", "Integration", "Medium", ["Gọi /api/v1/customer/restaurants", "Kiểm tra dữ liệu trả về"], "Mỗi phần tử có RestaurantID, RestaurantName, Opentime, Closetime"),
        tc("Xem thông tin nhà hàng không tồn tại", "RestaurantID không có trong DB", "Negative", "Medium", ["Gọi /api/v1/customer/restaurant/999999", "Quan sát phản hồi"], "API trả dữ liệu rỗng hoặc thông báo không tìm thấy, frontend không crash"),
        tc("Tìm kiếm không phân biệt hoa thường theo địa chỉ", "Có địa chỉ chứa từ khóa", "Boundary", "Low", ["Nhập từ khóa khác kiểu chữ hoa/thường", "Nhấn Tìm kiếm"], "Kết quả vẫn khớp do ilike"),
        tc("Tìm kiếm với khoảng trắng đầu cuối", "Có dữ liệu phù hợp", "Boundary", "Low", ["Nhập từ khóa có khoảng trắng đầu cuối", "Nhấn Tìm kiếm"], "Hệ thống vẫn tìm được hoặc xử lý khoảng trắng nhất quán"),
        tc("Không gọi API khi mất kết nối server", "Tắt server backend hoặc mô phỏng lỗi mạng", "Negative", "Medium", ["Nhấn Tìm kiếm", "Quan sát UI"], "Frontend hiển thị cảnh báo lỗi gọi API"),
        tc("Kết quả tìm kiếm hiển thị đủ thông tin chính", "Có ít nhất một kết quả", "UI", "Medium", ["Tìm kiếm nhà hàng", "Kiểm tra card kết quả"], "Card có tên, địa chỉ, số điện thoại, trạng thái bàn và hai nút thao tác"),
    ]),
}


def add_sheet_cases():
    SHEET_CASES.update({
        "CUS_BOOKING": ("CUS_BOOK", booking_cases()),
        "CUS_HIS": ("CUS_HIST", history_cases()),
        "RES_BOOKING": ("RES_BOOK", res_booking_cases()),
        "RES_MENU": ("RES_MENU", res_menu_cases()),
        "RES_TABLE": ("RES_TABLE", res_table_cases()),
        "RES_ORDER": ("RES_ORDER", res_order_cases()),
        "RES_PAYMENT": ("RES_PAY", res_payment_cases()),
        "ADMIN_USER": ("ADMIN_USER", admin_user_cases()),
        "ADMIN_CUS": ("ADMIN_CUS", admin_cuisine_cases()),
        "ADMIN_RES": ("ADMIN_RES", admin_restaurant_cases()),
        "ADMIN_REPORT": ("ADMIN_REPORT", admin_report_cases()),
    })


def booking_cases():
    pre = "Khách hàng ở trang booking.html, hệ thống có nhà hàng và bàn"
    return [
        tc("Đặt bàn thành công với thông tin hợp lệ", pre, "Function", "Critical", ["Chọn nhà hàng", "Nhập họ tên, SĐT, ngày, giờ, số người", "Kiểm tra bàn", "Chọn bàn trống"], "Tạo Reservation trạng thái Pending, trả reservation_id, QR và tiền cọc"),
        tc("Tính tiền cọc đúng theo số người", pre, "Function", "High", ["Nhập số người = 4", "Chọn bàn phù hợp", "Tạo đặt bàn"], "Deposit bằng 4 x 50.000 = 200.000 VNĐ"),
        tc("Kiểm tra bàn trống theo sức chứa", pre, "Function", "High", ["Nhập số người nhỏ hơn hoặc bằng Capacity", "Nhấn Kiểm tra bàn"], "Chỉ hiển thị bàn có Capacity >= số người"),
        tc("Không cho đặt khi bỏ trống họ tên", pre, "Validation", "High", ["Bỏ trống họ tên", "Nhập các trường còn lại", "Nhấn Kiểm tra bàn"], "Frontend/API báo tên là bắt buộc"),
        tc("Không cho đặt khi bỏ trống số điện thoại", pre, "Validation", "High", ["Bỏ trống SĐT", "Nhập các trường còn lại", "Chọn bàn"], "API /book trả 400 với thông báo tên và SĐT bắt buộc"),
        tc("Không cho kiểm tra bàn khi chưa chọn nhà hàng", pre, "Validation", "High", ["Không chọn nhà hàng", "Nhập thông tin đặt bàn", "Nhấn Kiểm tra bàn"], "Hiển thị yêu cầu chọn nhà hàng"),
        tc("Không cho kiểm tra bàn khi thiếu ngày", pre, "Validation", "Medium", ["Bỏ trống ngày", "Nhập các trường còn lại", "Nhấn Kiểm tra bàn"], "Hiển thị yêu cầu nhập đầy đủ thông tin"),
        tc("Không cho kiểm tra bàn khi thiếu giờ", pre, "Validation", "Medium", ["Bỏ trống giờ", "Nhập các trường còn lại", "Nhấn Kiểm tra bàn"], "Hiển thị yêu cầu nhập đầy đủ thông tin"),
        tc("Không cho đặt ngoài giờ mở cửa", pre, "Boundary", "High", ["Chọn nhà hàng", "Nhập giờ nhỏ hơn Opentime", "Nhấn Kiểm tra bàn"], "Frontend chặn và thông báo chỉ được đặt trong giờ hoạt động"),
        tc("Cho đặt đúng bằng giờ mở cửa", pre, "Boundary", "Medium", ["Nhập giờ bằng Opentime", "Nhấn Kiểm tra bàn"], "Hệ thống chấp nhận kiểm tra bàn"),
        tc("Cho đặt đúng bằng giờ đóng cửa", pre, "Boundary", "Medium", ["Nhập giờ bằng Closetime", "Nhấn Kiểm tra bàn"], "Hệ thống chấp nhận theo rule hiện tại của frontend"),
        tc("Không cho double booking cùng bàn cùng ngày giờ", "Đã có Reservation Pending/Confirmed cùng TableID, ngày, giờ", "Negative", "Critical", ["Chọn cùng bàn và cùng khung giờ", "Gửi tạo đặt bàn"], "API trả lỗi bàn đã được đặt trong khung giờ này"),
        tc("Không cho đặt table_id không thuộc nhà hàng đã chọn", pre, "Security", "High", ["Gửi API /book với restaurant_id A và table_id của nhà hàng B"], "API trả 400 bàn không hợp lệ"),
        tc("Không cho đặt với table_id không tồn tại", pre, "Negative", "High", ["Gửi /book với table_id = 999999", "Quan sát phản hồi"], "API trả 400 bàn không hợp lệ"),
        tc("Kiểm tra định dạng ngày sai", pre, "Negative", "Medium", ["Gọi /check với date sai định dạng", "Quan sát phản hồi"], "Hệ thống không tạo booking và trả lỗi xử lý dữ liệu"),
        tc("Kiểm tra số người bằng 0", pre, "Boundary", "High", ["Nhập số người = 0", "Nhấn Kiểm tra bàn"], "Hệ thống cần chặn số người không hợp lệ, không tạo tiền cọc 0"),
        tc("Kiểm tra số người âm", pre, "Boundary", "High", ["Nhập số người = -1", "Nhấn Kiểm tra bàn"], "Hệ thống cần chặn dữ liệu âm, không hiển thị bàn"),
        tc("Đặt bàn khách vãng lai không có token", pre, "Function", "Medium", ["Không đăng nhập", "Nhập thông tin hợp lệ", "Chọn bàn"], "Booking vẫn được tạo với UserID rỗng và trạng thái Pending"),
        tc("Đặt bàn khách đã đăng nhập", "Token CUSTOMER hợp lệ", "Function", "High", ["Đăng nhập", "Đặt bàn hợp lệ", "Kiểm tra lịch sử"], "Booking lưu UserID của token và xuất hiện trong lịch sử khách hàng"),
        tc("Tự hủy booking Pending quá hạn", "Có booking Pending cũ hơn thời gian quy định", "Integration", "Medium", ["Gọi tạo booking mới để kích hoạt cancel_expired_bookings", "Kiểm tra booking cũ"], "Booking Pending quá hạn chuyển sang Cancelled"),
    ]


def history_cases():
    pre = "Khách hàng đã đăng nhập và có dữ liệu Reservations"
    return [
        tc("Xem lịch sử đặt bàn thành công", pre, "Function", "High", ["Mở history.html", "Frontend gọi /customer/history"], "Hiển thị danh sách booking của user hiện tại"),
        tc("Chặn xem lịch sử khi chưa đăng nhập", "Không có token trong localStorage", "Security", "Critical", ["Mở history.html"], "Chuyển về trang đăng nhập và không gọi dữ liệu riêng tư"),
        tc("Chặn token không hợp lệ khi xem lịch sử", "Token bị sửa hoặc hết hạn", "Security", "High", ["Gọi /customer/me", "Gọi /customer/history"], "API trả 401 và frontend yêu cầu đăng nhập lại"),
        tc("Tìm lịch sử theo tên khách hàng", pre, "Function", "Medium", ["Nhập tên vào ô tìm kiếm", "Quan sát danh sách"], "Chỉ hiển thị booking có CustomerName chứa từ khóa"),
        tc("Tìm lịch sử theo mã ReservationID", pre, "Function", "Medium", ["Nhập mã đơn dạng số", "Quan sát danh sách"], "Hiển thị đúng đơn có ReservationID tương ứng"),
        tc("Tìm lịch sử không có kết quả", pre, "Negative", "Low", ["Nhập từ khóa không tồn tại"], "Hiển thị thông báo không có dữ liệu"),
        tc("Lọc lịch sử trạng thái Pending", pre, "Function", "Medium", ["Chọn filter Pending"], "Chỉ hiển thị đơn Pending"),
        tc("Lọc lịch sử trạng thái Confirmed", pre, "Function", "Medium", ["Chọn filter Confirmed"], "Chỉ hiển thị đơn Confirmed"),
        tc("Lọc lịch sử trạng thái Cancelled", pre, "Function", "Medium", ["Chọn filter Cancelled"], "Chỉ hiển thị đơn Cancelled"),
        tc("Kết hợp tìm kiếm và lọc trạng thái", pre, "Function", "Medium", ["Nhập từ khóa", "Chọn trạng thái"], "Danh sách thỏa đồng thời từ khóa và trạng thái"),
        tc("Không hiển thị booking của user khác", "Có booking thuộc nhiều user", "Security", "Critical", ["Đăng nhập user A", "Mở lịch sử"], "Chỉ trả Reservation.UserID của user A"),
        tc("Hiển thị thông tin nhà hàng trong lịch sử", pre, "Function", "Medium", ["Mở lịch sử", "Kiểm tra từng card"], "Có RestaurantName, RestaurantID nếu dữ liệu nhà hàng tồn tại"),
        tc("Sắp xếp lịch sử mới nhất lên đầu", pre, "Function", "Low", ["Mở lịch sử", "So sánh mã ReservationID"], "Danh sách order by ReservationID desc"),
        tc("Gửi đánh giá lần đầu cho booking hợp lệ", "Có booking đủ điều kiện đánh giá", "Function", "High", ["Mở popup đánh giá", "Nhập rating và comment", "Nhấn Gửi"], "API /review tạo Review và trả thông báo thành công"),
        tc("Cập nhật đánh giá đã tồn tại", "Booking đã có Review", "Function", "Medium", ["Mở lại đánh giá", "Sửa rating/comment", "Nhấn Gửi"], "Review cũ được cập nhật, không tạo trùng ReservationID"),
        tc("Kiểm tra review đã tồn tại", "Có Review theo ReservationID", "Function", "Medium", ["Gọi /review/check?reservation_id={id}"], "API trả reviewed=true, rating và comment"),
        tc("Không gửi review khi thiếu ReservationID", pre, "Validation", "High", ["Mở popup thiếu mã đơn", "Nhấn Gửi"], "Frontend cảnh báo thiếu mã đặt bàn"),
        tc("Review với rating ngoài khoảng 1-5", "Có booking để đánh giá", "Boundary", "Medium", ["Nhập rating = 6 hoặc 0", "Nhấn Gửi"], "Hệ thống cần chặn hoặc không lưu giá trị ngoài khoảng"),
        tc("Review khi chưa đăng nhập", "Không có token", "Security", "High", ["Gọi POST /customer/review"], "API trả 401 chưa đăng nhập"),
        tc("Nút đóng lịch sử quay về trang chủ", pre, "UI", "Low", ["Nhấn Đóng"], "Điều hướng về home.html"),
    ]


def res_booking_cases():
    pre = "Nhân viên nhà hàng đã đăng nhập bằng role STAFF"
    return [
        tc("Xem danh sách booking của nhà hàng", pre, "Function", "High", ["Mở comfirmbooking.html", "Gọi /restaurant/bookings"], "Chỉ hiển thị booking thuộc RestaurantID của staff"),
        tc("Xác nhận booking Pending thành công", "Có booking Pending", "Function", "Critical", ["Chọn booking Pending", "Nhấn Xác nhận"], "Reservation.Status chuyển Confirmed và bàn chuyển Reserved"),
        tc("Từ chối booking Pending thành công", "Có booking Pending", "Function", "High", ["Chọn booking Pending", "Nhấn Từ chối"], "Reservation.Status chuyển Rejected"),
        tc("Xóa booking thành công", "Có booking cần xóa", "Function", "Medium", ["Chọn booking", "Nhấn Xóa", "Xác nhận"], "Booking bị xóa khỏi DB và danh sách cập nhật"),
        tc("Xác nhận booking không tồn tại", pre, "Negative", "High", ["Gọi /bookings/999999/confirm"], "API trả lỗi Booking not found"),
        tc("Từ chối booking không tồn tại", pre, "Negative", "Medium", ["Gọi /bookings/999999/reject"], "API trả lỗi Not found"),
        tc("Xóa booking không tồn tại", pre, "Negative", "Medium", ["Gọi DELETE /bookings/999999"], "API trả lỗi Not found"),
        tc("Tạo booking thủ công từ phía nhà hàng", pre, "Function", "High", ["Gửi POST /restaurant/bookings với BookingDate, BookingTime, TableID", "Quan sát phản hồi"], "Booking được tạo Pending và bàn chuyển Reserved"),
        tc("Tạo booking thủ công sai định dạng ngày giờ", pre, "Validation", "High", ["Gửi BookingDate/BookingTime sai format", "Quan sát phản hồi"], "API trả Wrong datetime format"),
        tc("Tạo booking thủ công với bàn không tồn tại", pre, "Negative", "High", ["Gửi TableID không tồn tại", "Quan sát phản hồi"], "API trả Table not found"),
        tc("Xem booking theo bàn", "Có TableID hợp lệ", "Function", "Medium", ["Gọi /tables/{id}/bookings"], "Trả danh sách Reservation của bàn đó"),
        tc("Danh sách booking sắp xếp theo giờ đặt", pre, "Function", "Low", ["Tải danh sách booking", "Quan sát thứ tự"], "Dữ liệu order theo BookingTime giảm dần"),
        tc("Không cho CUSTOMER xác nhận booking", "Đăng nhập role CUSTOMER", "Security", "High", ["Gọi /restaurant/bookings/{id}/confirm bằng token CUSTOMER"], "Hệ thống cần chặn quyền hoặc không cho thao tác nhà hàng"),
        tc("Không cho xem booking khi thiếu token", "Không có token", "Security", "High", ["Gọi GET /restaurant/bookings"], "API trả 401"),
        tc("Hiển thị đầy đủ thông tin booking", "Có dữ liệu booking", "UI", "Medium", ["Mở màn quản lý booking"], "Có tên khách, SĐT, ngày, giờ, số khách, bàn, trạng thái"),
        tc("Xác nhận nhiều booking khác bàn cùng giờ", "Có hai booking ở hai bàn khác nhau", "Integration", "Medium", ["Xác nhận từng booking"], "Cả hai booking được xác nhận độc lập"),
        tc("Từ chối booking không làm đổi trạng thái bàn đang Reserved bởi booking khác", "Bàn có booking khác Confirmed", "Integration", "Medium", ["Từ chối booking Pending cùng bàn khác giờ"], "Không phá vỡ trạng thái bàn của booking đã xác nhận"),
        tc("Cập nhật trạng thái bàn từ màn booking", "Có TableID hợp lệ", "Function", "Medium", ["Gọi PUT /tables/{id}/status", "Gửi Status mới"], "Bàn cập nhật trạng thái theo request"),
        tc("Cập nhật trạng thái bàn không tồn tại", pre, "Negative", "Medium", ["Gọi PUT /tables/999999/status"], "API trả Table not found"),
        tc("Load booking khi nhà hàng chưa có đơn", "RestaurantID của staff không có Reservation", "Boundary", "Low", ["Mở màn booking"], "Hiển thị danh sách rỗng, không lỗi giao diện"),
    ]


def res_menu_cases():
    pre = "Nhân viên STAFF đã đăng nhập và được gán RestaurantID"
    return [
        tc("Xem danh sách menu quản trị", pre, "Function", "High", ["Mở menu_mgmt.html", "Gọi /restaurant/menu/admin"], "Hiển thị món thuộc đúng nhà hàng của staff"),
        tc("Thêm món mới thành công", pre, "Function", "Critical", ["Nhập tên món, giá, ảnh, category", "Nhấn Thêm món"], "Tạo Food mới, trả id và món xuất hiện trong danh sách"),
        tc("Không thêm món khi bỏ trống tên", pre, "Validation", "High", ["Bỏ trống tên món", "Nhập giá", "Nhấn Thêm món"], "Frontend cảnh báo cần nhập tên và giá"),
        tc("Không thêm món khi bỏ trống giá", pre, "Validation", "High", ["Nhập tên món", "Bỏ trống giá", "Nhấn Thêm món"], "Frontend cảnh báo cần nhập tên và giá"),
        tc("Thêm món với giá bằng 0", pre, "Boundary", "Medium", ["Nhập giá = 0", "Nhấn Thêm món"], "Hệ thống cần từ chối giá không hợp lệ hoặc không lưu món giá 0"),
        tc("Thêm món với giá âm", pre, "Boundary", "High", ["Nhập giá = -10000", "Nhấn Thêm món"], "Hệ thống cần chặn giá âm"),
        tc("Sửa tên và giá món thành công", "Có món trong danh sách", "Function", "High", ["Nhấn Sửa", "Nhập tên mới và giá mới", "Lưu"], "Food cập nhật tên/giá và giao diện load lại"),
        tc("Sửa món với tên rỗng", "Có món trong danh sách", "Validation", "Medium", ["Nhấn Sửa", "Xóa tên món", "Lưu"], "Frontend cảnh báo dữ liệu không hợp lệ"),
        tc("Sửa món với giá không phải số", "Có món trong danh sách", "Validation", "Medium", ["Nhấn Sửa", "Nhập giá abc", "Lưu"], "Frontend cảnh báo dữ liệu không hợp lệ"),
        tc("Upload ảnh mới khi sửa món", "Có file ảnh hợp lệ", "Function", "Medium", ["Nhấn Sửa", "Chọn ảnh mới", "Lưu"], "File được lưu và Image_URL cập nhật"),
        tc("Sửa món không chọn ảnh mới", "Có món có ảnh cũ", "Function", "Low", ["Nhấn Sửa", "Chọn Cancel khi hỏi ảnh", "Lưu tên/giá"], "Món giữ nguyên ảnh cũ"),
        tc("Xóa món thành công", "Có món cần xóa", "Function", "High", ["Nhấn Xóa", "Xác nhận"], "Food bị xóa và không còn trong danh sách"),
        tc("Xóa món không tồn tại", pre, "Negative", "Medium", ["Gọi DELETE /menu/xxxxx", "Quan sát phản hồi"], "API trả Food not found"),
        tc("Ẩn món khỏi menu order", "Có món Visible=true", "Function", "High", ["Nhấn Ẩn món", "Mở màn order"], "Visible=false, món không hiển thị trong menu order"),
        tc("Hiện lại món đã ẩn", "Có món Visible=false", "Function", "High", ["Nhấn Hiện món", "Mở màn order"], "Visible=true, món hiển thị lại"),
        tc("Toggle món không tồn tại", pre, "Negative", "Medium", ["Gọi PUT /menu/xxxxx/toggle"], "API trả 404 Not found"),
        tc("Không cho xem menu admin khi role không phải STAFF", "Đăng nhập CUSTOMER hoặc ADMIN", "Security", "High", ["Gọi /restaurant/menu/admin"], "API trả 403 chỉ nhân viên mới được xem"),
        tc("Không cho thêm món khi staff chưa gán nhà hàng", "User STAFF có RestaurantID rỗng", "Security", "High", ["Gọi POST /restaurant/menu"], "API trả nhân viên chưa được phân công nhà hàng"),
        tc("Menu order chỉ lấy món Visible=true", pre, "Integration", "Medium", ["Ẩn một món", "Gọi GET /restaurant/menu"], "API không trả món đã ẩn"),
        tc("Hiển thị ảnh mặc định khi món thiếu ảnh", "Có món Image_URL rỗng", "UI", "Low", ["Mở menu_mgmt.html", "Quan sát món"], "Giao diện dùng ảnh mặc định hoặc ảnh từ MENU_DATA"),
    ]


def res_table_cases():
    pre = "Nhân viên STAFF đã đăng nhập và được gán RestaurantID"
    return [
        tc("Xem danh sách bàn của nhà hàng", pre, "Function", "High", ["Mở table.html", "Gọi /restaurant/tables"], "Chỉ hiển thị bàn thuộc RestaurantID của staff"),
        tc("Thêm bàn mới thành công", pre, "Function", "Critical", ["Nhấn Thêm bàn", "Nhập TableNumber và Capacity", "Lưu"], "Tạo RestaurantTables và hiển thị trong danh sách"),
        tc("Không thêm bàn trùng TableNumber", "Đã có bàn cùng TableNumber", "Negative", "High", ["Thêm bàn với số bàn trùng", "Lưu"], "API trả Table already exists"),
        tc("Thêm bàn khi bỏ trống Capacity", pre, "Boundary", "Medium", ["Nhập TableNumber", "Bỏ trống Capacity", "Lưu"], "Hệ thống dùng Capacity mặc định 4 theo service"),
        tc("Thêm bàn với Capacity bằng 1", pre, "Boundary", "Low", ["Nhập Capacity = 1", "Lưu"], "Bàn được tạo nếu rule cho phép"),
        tc("Thêm bàn với Capacity bằng 0", pre, "Boundary", "High", ["Nhập Capacity = 0", "Lưu"], "Hệ thống cần chặn sức chứa không hợp lệ hoặc không lưu 0"),
        tc("Thêm bàn với Capacity âm", pre, "Boundary", "High", ["Nhập Capacity = -4", "Lưu"], "Hệ thống cần chặn dữ liệu âm"),
        tc("Thêm bàn với Capacity không phải số", pre, "Validation", "High", ["Nhập Capacity = abc", "Lưu"], "Hệ thống không tạo bàn và không lỗi 500"),
        tc("Cập nhật trạng thái bàn sang Reserved", "Có bàn hợp lệ", "Function", "High", ["Gọi PUT /tables/{id}/status", "Gửi Status=Reserved"], "Bàn chuyển trạng thái Reserved"),
        tc("Cập nhật trạng thái bàn sang Trống", "Có bàn hợp lệ", "Function", "High", ["Gọi PUT /tables/{id}/status", "Gửi Status=Trống"], "Bàn chuyển trạng thái Trống"),
        tc("Cập nhật trạng thái bàn không tồn tại", pre, "Negative", "Medium", ["Gọi PUT /tables/999999/status"], "API trả Table not found"),
        tc("Hiển thị khách đang dùng bàn đã Confirmed", "Có Reservation Confirmed theo TableID", "Function", "Medium", ["Mở table.html", "Quan sát bàn"], "Bàn hiển thị customer_name và customer_count"),
        tc("Bàn không có booking không hiển thị khách", "Bàn chưa có Reservation Confirmed", "Function", "Low", ["Mở table.html"], "customer_name rỗng và customer_count = 0"),
        tc("Không cho xem danh sách bàn khi thiếu token", "Không có token", "Security", "High", ["Gọi GET /restaurant/tables"], "API trả 401"),
        tc("Không cho tạo bàn khi thiếu token", "Không có token", "Security", "High", ["Gọi POST /restaurant/tables"], "API trả 401"),
        tc("Staff nhà hàng A không thấy bàn nhà hàng B", "Có dữ liệu nhiều nhà hàng", "Security", "Critical", ["Đăng nhập staff A", "Tải danh sách bàn"], "Chỉ trả bàn có RestaurantID của staff A"),
        tc("Điều hướng từ bàn sang màn order", "Có bàn trong danh sách", "Function", "Medium", ["Nhấn thao tác order trên bàn", "Quan sát URL"], "Chuyển sang orders.html?table_id={id}"),
        tc("Danh sách bàn hiển thị hai cột ổn định", pre, "UI", "Low", ["Mở table.html với nhiều bàn"], "Bàn được phân bổ trên giao diện, không vỡ layout"),
        tc("Bàn Reserved sau khi xác nhận booking", "Có booking Pending của bàn", "Integration", "High", ["Xác nhận booking", "Mở lại table.html"], "Trạng thái bàn hiển thị Reserved"),
        tc("Bàn Trống sau khi thanh toán order", "Có order active theo bàn", "Integration", "High", ["Thanh toán order", "Mở lại table.html"], "Trạng thái bàn trở về Trống"),
    ]


def res_order_cases():
    pre = "Nhân viên ở màn orders.html với table_id hợp lệ"
    return [
        tc("Tải menu order thành công", pre, "Function", "High", ["Mở orders.html?table_id=1", "Gọi GET /restaurant/menu"], "Hiển thị các món Visible=true của nhà hàng"),
        tc("Thêm một món vào order", pre, "Function", "High", ["Nhấn Thêm ở một món"], "Món xuất hiện trong danh sách order với số lượng 1"),
        tc("Tăng số lượng khi thêm cùng món nhiều lần", pre, "Function", "Medium", ["Nhấn Thêm cùng món 3 lần"], "Danh sách hiển thị món x3 và tổng tiền cập nhật"),
        tc("Tính tổng tiền nhiều món", pre, "Function", "High", ["Thêm nhiều món giá khác nhau"], "Tổng tiền bằng tổng price x qty"),
        tc("Chặn gửi order khi chưa có món", pre, "Validation", "High", ["Không chọn món", "Nhấn Gửi order"], "Frontend cảnh báo chưa có món"),
        tc("Gửi order mới thành công", pre, "Function", "Critical", ["Chọn món", "Nhấn Gửi order"], "API /orders tạo order active và order_item tương ứng"),
        tc("Gửi thêm món vào order active đã có", "Bàn đã có order active", "Function", "High", ["Chọn thêm món", "Nhấn Gửi order"], "Order cũ được cập nhật, không tạo order active trùng"),
        tc("Gộp số lượng món đã tồn tại trong order", "Order active đã có cùng food_id", "Function", "High", ["Gửi cùng food_id với qty mới"], "OrderItem.quantity được cộng thêm"),
        tc("Gửi order với table_id không tồn tại", "URL có table_id không tồn tại", "Negative", "High", ["Gửi order"], "API trả 404 Table not found"),
        tc("Gửi order thiếu table_id", "URL không có table_id", "Validation", "High", ["Mở orders.html không có table_id", "Chọn món và gửi"], "Hệ thống cần chặn hoặc báo lỗi dữ liệu không hợp lệ"),
        tc("Gửi order với qty bằng 0", pre, "Boundary", "Medium", ["Sửa payload qty=0", "Gửi API"], "Hệ thống cần không lưu item số lượng 0"),
        tc("Gửi order với qty âm", pre, "Boundary", "High", ["Sửa payload qty=-1", "Gửi API"], "Hệ thống cần từ chối dữ liệu âm"),
        tc("Gửi order với food_id không tồn tại", pre, "Negative", "Medium", ["Sửa payload food_id sai", "Gửi API"], "Hệ thống cần báo lỗi hoặc không tạo OrderItem mồ côi"),
        tc("Chỉ hiển thị món đang Visible", "Có món Visible=false", "Integration", "Medium", ["Mở orders.html", "Quan sát menu"], "Món bị ẩn không xuất hiện"),
        tc("Ảnh món lỗi hiển thị placeholder", "Có món image URL lỗi", "UI", "Low", ["Mở orders.html", "Quan sát ảnh món"], "Ảnh fallback placeholder được hiển thị"),
        tc("Khôi phục order tạm trong localStorage", "localStorage có currentOrder cùng table_id", "Function", "Low", ["Mở orders.html", "Quan sát order-list"], "Order tạm được render lại"),
        tc("Không khôi phục order tạm khác table_id", "localStorage có currentOrder của bàn khác", "Function", "Low", ["Mở bàn khác", "Quan sát order-list"], "Không hiển thị order của bàn khác"),
        tc("Gửi order khi backend mất kết nối", pre, "Negative", "Medium", ["Tắt backend", "Nhấn Gửi order"], "Frontend cảnh báo lỗi gửi order"),
        tc("Order làm bàn chuyển Reserved", "Bàn đang Trống", "Integration", "Medium", ["Gửi order mới", "Kiểm tra RestaurantTables"], "Bàn được cập nhật trạng thái Reserved"),
        tc("Không cho CUSTOMER thao tác order nhà hàng", "Đăng nhập CUSTOMER", "Security", "High", ["Mở hoặc gọi API order"], "Hệ thống cần chặn truy cập chức năng nội bộ nhà hàng"),
    ]


def res_payment_cases():
    pre = "Bàn có order active trong hệ thống"
    return [
        tc("Thanh toán order active thành công", pre, "Function", "Critical", ["Mở orders.html của bàn", "Nhấn Thanh toán", "Xác nhận"], "Order chuyển paid và bàn chuyển Trống"),
        tc("Thanh toán khi chưa có order active", "Bàn hợp lệ nhưng không có order active", "Boundary", "Medium", ["Gọi PUT /orders/pay/{table_id}"], "Service tạo order paid theo logic hiện tại và trả 200"),
        tc("Thanh toán table_id không tồn tại", "Không có bàn theo id", "Negative", "High", ["Gọi PUT /orders/pay/999999"], "API không làm lỗi server; cần báo bàn không tồn tại hoặc xử lý rõ ràng"),
        tc("Hủy xác nhận thanh toán trên UI", pre, "Function", "Medium", ["Nhấn Thanh toán", "Chọn Cancel ở hộp xác nhận"], "Không gọi thanh toán và vẫn ở màn order"),
        tc("Xóa currentOrder sau thanh toán", pre, "Function", "Medium", ["Tạo currentOrder", "Thanh toán thành công"], "localStorage currentOrder bị xóa"),
        tc("Điều hướng về quản lý bàn sau thanh toán", pre, "UI", "Low", ["Thanh toán thành công"], "Chuyển về table.html"),
        tc("Thanh toán nhiều lần cùng một bàn", "Order đã paid", "Boundary", "Medium", ["Gọi pay lần 1", "Gọi pay lần 2"], "Không tạo lỗi hệ thống; trạng thái cuối vẫn paid/Trống"),
        tc("Bàn trở về Trống sau thanh toán", pre, "Integration", "High", ["Thanh toán order", "Gọi GET /restaurant/tables"], "Status bàn là Trống"),
        tc("Order đổi từ active sang paid", pre, "Integration", "High", ["Thanh toán order", "Kiểm tra bảng orders"], "orders.status = paid"),
        tc("Thanh toán khi mất kết nối backend", pre, "Negative", "Medium", ["Tắt backend", "Nhấn Thanh toán"], "Frontend không báo thành công giả khi API thất bại"),
        tc("Thanh toán đặt cọc booking thành công", "Có Reservation Pending và amount đúng Deposit", "Function", "Critical", ["Gọi POST /customer/payment", "Gửi reservation_id và amount đúng"], "Tạo Payment Status=Paid, PaymentMethod=QR"),
        tc("Thanh toán đặt cọc sai số tiền", "Có Reservation Pending", "Negative", "High", ["Gửi amount khác Deposit"], "API trả lỗi Sai số tiền và không tạo Payment"),
        tc("Thanh toán đặt cọc reservation không tồn tại", "Không có ReservationID", "Negative", "High", ["Gửi reservation_id=999999"], "API trả Not found"),
        tc("Thanh toán đặt cọc booking đã xử lý", "Reservation không còn Pending", "Negative", "High", ["Gửi payment cho booking Confirmed/Rejected"], "API trả Booking already processed"),
        tc("Thanh toán đặt cọc thiếu amount", "Có Reservation Pending", "Validation", "High", ["Gửi payment không có amount"], "API không tạo Payment và trả lỗi dữ liệu"),
        tc("Thanh toán đặt cọc thiếu reservation_id", "Người dùng ở popup thanh toán", "Validation", "High", ["Gửi payment không có reservation_id"], "API trả Not found hoặc lỗi dữ liệu phù hợp"),
        tc("QR đặt cọc chứa mã ReservationID", "Đặt bàn thành công", "Function", "Medium", ["Tạo booking", "Kiểm tra URL QR trả về"], "QR chứa addInfo DATBAN_{reservation_id}"),
        tc("Số tiền cọc hiển thị đúng trên popup", "Đặt bàn thành công", "UI", "Medium", ["Mở popup thanh toán", "So sánh deposit"], "Tiền cọc hiển thị bằng deposit backend trả về"),
        tc("Không xác nhận thanh toán khi response có error", "API payment trả error", "Negative", "Medium", ["Nhấn Đã chuyển khoản", "Quan sát popup"], "Hiển thị lỗi và không mở popup thành công"),
        tc("Báo cáo admin ghi nhận payment đã tạo", "Có Payment CreatedAt trong tháng", "Integration", "Medium", ["Thanh toán đặt cọc", "Mở báo cáo tháng"], "Doanh thu nhà hàng tăng theo Amount"),
    ]


def admin_user_cases():
    pre = "Admin đã đăng nhập và có token role ADMIN"
    return [
        tc("Xem danh sách user thành công", pre, "Function", "Critical", ["Mở users.html", "Gọi GET /admin/users"], "Hiển thị id, username, email, phone, role"),
        tc("Chặn xem user khi không phải ADMIN", "Đăng nhập CUSTOMER hoặc STAFF", "Security", "Critical", ["Gọi GET /admin/users"], "API trả 403"),
        tc("Chặn xem user khi thiếu token", "Không có token", "Security", "High", ["Gọi GET /admin/users"], "API trả 401"),
        tc("Cập nhật thông tin user thành công", pre, "Function", "High", ["Chọn user", "Sửa Username, Email, Phone, Role", "Lưu"], "API trả cập nhật thành công và danh sách reload"),
        tc("Cập nhật role user sang STAFF", pre, "Function", "High", ["Mở modal sửa user", "Chọn Role STAFF", "Lưu"], "Role trong DB đổi thành STAFF"),
        tc("Cập nhật role user sang ADMIN", pre, "Security", "High", ["Mở modal sửa user", "Chọn Role ADMIN", "Lưu"], "Role đổi thành ADMIN theo quyền admin"),
        tc("Cập nhật user không tồn tại", pre, "Negative", "Medium", ["Gọi PUT /users/999999", "Quan sát phản hồi"], "API trả 404 không tìm thấy người dùng"),
        tc("Cập nhật email sai định dạng", pre, "Validation", "Medium", ["Nhập email abc@", "Lưu"], "Hệ thống cần chặn hoặc không lưu email sai định dạng"),
        tc("Cập nhật phone quá 11 ký tự", pre, "Boundary", "Medium", ["Nhập phone 12 chữ số", "Lưu"], "Hệ thống cần chặn hoặc không lưu quá độ dài"),
        tc("Cập nhật username trống", pre, "Validation", "High", ["Xóa username", "Lưu"], "Hệ thống cần chặn dữ liệu bắt buộc"),
        tc("Xóa user thành công", "Có user không phải tài khoản đang dùng", "Function", "High", ["Nhấn Xóa", "Xác nhận"], "User bị xóa và danh sách cập nhật"),
        tc("Xóa user không tồn tại", pre, "Negative", "Medium", ["Gọi DELETE /users/999999"], "API trả lỗi khi xóa người dùng"),
        tc("Chặn xóa user khi không phải ADMIN", "Đăng nhập STAFF", "Security", "High", ["Gọi DELETE /admin/users/{id}"], "API trả 403"),
        tc("Không xác định ID khi lưu modal thêm mới", pre, "UI", "Low", ["Mở modal thêm user", "Nhấn Lưu"], "Frontend cảnh báo không xác định user cần sửa"),
        tc("Đóng modal sửa user", pre, "UI", "Low", ["Mở modal sửa user", "Nhấn Hủy"], "Modal đóng, dữ liệu bảng không đổi"),
        tc("Danh sách user hiển thị phone rỗng", "Có user chưa có phone", "UI", "Low", ["Mở users.html"], "Hiển thị Chưa có số hoặc giá trị phù hợp"),
        tc("Không hiển thị password trong danh sách", pre, "Security", "Critical", ["Gọi GET /admin/users", "Kiểm tra JSON"], "Response không chứa Password"),
        tc("Cập nhật chỉ một trường user", pre, "Function", "Medium", ["Chỉ sửa Phone", "Lưu"], "Chỉ Phone thay đổi, các trường khác giữ nguyên"),
        tc("Reload danh sách sau cập nhật", pre, "Integration", "Low", ["Cập nhật user thành công", "Quan sát bảng"], "fetchUsers được gọi lại và dữ liệu mới hiển thị"),
        tc("Đăng xuất khỏi màn quản lý user", pre, "Function", "Low", ["Nhấn Đăng xuất", "Xác nhận"], "Token bị xóa và chuyển về login.html"),
    ]


def admin_cuisine_cases():
    pre = "Admin đã đăng nhập và đang ở cuisine.html"
    return [
        tc("Xem danh sách cuisine thành công", pre, "Function", "High", ["Mở cuisine.html", "Gọi GET /admin/cuisines"], "Hiển thị id, name, status"),
        tc("Thêm cuisine mới thành công", pre, "Function", "Critical", ["Nhấn Thêm mới", "Nhập CuisineName chưa tồn tại", "Lưu"], "API trả 201 và cuisine xuất hiện trong danh sách"),
        tc("Không thêm cuisine khi bỏ trống tên", pre, "Validation", "High", ["Mở modal thêm", "Bỏ trống tên", "Lưu"], "API trả 400 tên không được để trống"),
        tc("Không thêm cuisine trùng tên", "CuisineName đã tồn tại", "Negative", "High", ["Nhập tên trùng", "Lưu"], "API trả 400 danh mục đã tồn tại"),
        tc("Thêm cuisine với khoảng trắng đầu cuối", pre, "Boundary", "Medium", ["Nhập tên có khoảng trắng", "Lưu"], "Hệ thống cần trim hoặc xử lý nhất quán, không tạo bản ghi rác"),
        tc("Thêm cuisine tên rất dài", pre, "Boundary", "Medium", ["Nhập tên trên 100 ký tự", "Lưu"], "Hệ thống cần chặn quá độ dài hoặc trả lỗi DB rõ ràng"),
        tc("Cập nhật tên cuisine thành công", "Có cuisine hiện có", "Function", "High", ["Nhấn Sửa", "Nhập tên mới", "Lưu"], "API trả cập nhật thành công"),
        tc("Cập nhật trạng thái cuisine", "Có cuisine hiện có", "Function", "Medium", ["Nhấn Sửa", "Đổi Status", "Lưu"], "Status cập nhật theo lựa chọn"),
        tc("Cập nhật cuisine không tồn tại", pre, "Negative", "Medium", ["Gọi PUT /cuisines/999999"], "API trả 404 không tìm thấy"),
        tc("Cập nhật cuisine với tên trống", "Có cuisine hiện có", "Validation", "High", ["Sửa CuisineName thành rỗng", "Lưu"], "Hệ thống cần chặn tên trống"),
        tc("Xóa cuisine thành công", "Có cuisine không bị ràng buộc bởi nhà hàng", "Function", "High", ["Nhấn Xóa", "Xác nhận"], "Cuisine bị xóa khỏi danh sách"),
        tc("Xóa cuisine không tồn tại", pre, "Negative", "Medium", ["Gọi DELETE /cuisines/999999"], "API trả 404 không tìm thấy"),
        tc("Chặn thêm cuisine khi không phải ADMIN", "Đăng nhập STAFF", "Security", "Critical", ["Gọi POST /admin/cuisines"], "API trả 403"),
        tc("Chặn sửa cuisine khi không phải ADMIN", "Đăng nhập CUSTOMER", "Security", "High", ["Gọi PUT /admin/cuisines/{id}"], "API trả 403"),
        tc("Chặn xóa cuisine khi không phải ADMIN", "Đăng nhập STAFF", "Security", "High", ["Gọi DELETE /admin/cuisines/{id}"], "API trả 403"),
        tc("Load cuisine khi thiếu token", "Không có token", "Security", "High", ["Gọi GET /admin/cuisines"], "API yêu cầu JWT hợp lệ"),
        tc("Modal thêm cuisine reset dữ liệu cũ", pre, "UI", "Low", ["Mở sửa một cuisine", "Đóng modal", "Mở thêm mới"], "Ô tên được làm rỗng"),
        tc("Đóng modal cuisine không lưu", pre, "UI", "Low", ["Mở modal", "Nhập dữ liệu", "Nhấn Hủy"], "Không gọi API và dữ liệu không đổi"),
        tc("Danh sách cuisine reload sau xóa", "Có cuisine để xóa", "Integration", "Low", ["Xóa cuisine thành công"], "fetchCuisines được gọi lại, dòng bị xóa biến mất"),
        tc("Cuisine được dùng làm bộ lọc tìm kiếm", "Có cuisine đã tạo", "Integration", "Medium", ["Tạo cuisine", "Gán cho nhà hàng", "Tìm kiếm theo cuisine"], "Nhà hàng được lọc đúng theo CuisineID"),
    ]


def admin_restaurant_cases():
    pre = "Admin đã đăng nhập và đang ở restaurants.html"
    return [
        tc("Xem danh sách nhà hàng thành công", pre, "Function", "Critical", ["Mở restaurants.html", "Gọi GET /admin/restaurants"], "Hiển thị đầy đủ nhà hàng theo RestaurantID tăng dần"),
        tc("Lọc nhà hàng theo trạng thái", pre, "Function", "Medium", ["Gọi /admin/restaurants?status=Đang chờ duyệt"], "Chỉ trả nhà hàng có status tương ứng"),
        tc("Thêm nhà hàng bởi admin thành công", pre, "Function", "Critical", ["Nhấn Thêm nhà hàng", "Nhập tên và thông tin hợp lệ", "Lưu"], "Restaurant được tạo với status Đang hoạt động"),
        tc("Không thêm nhà hàng khi bỏ trống tên", pre, "Validation", "High", ["Bỏ trống RestaurantName", "Lưu"], "API trả 400 tên nhà hàng không được để trống"),
        tc("Thêm nhà hàng với email sai định dạng", pre, "Validation", "Medium", ["Nhập email abc@", "Lưu"], "Hệ thống cần chặn hoặc không lưu email sai định dạng"),
        tc("Thêm nhà hàng với phone quá 11 ký tự", pre, "Boundary", "Medium", ["Nhập phone 12 chữ số", "Lưu"], "Hệ thống cần chặn hoặc không lưu quá độ dài"),
        tc("Thêm nhà hàng giờ mở cửa sau giờ đóng cửa", pre, "Boundary", "High", ["Nhập Opentime 22:00, Closetime 08:00", "Lưu"], "Hệ thống cần chặn rule giờ không hợp lệ"),
        tc("Cập nhật nhà hàng thành công", "Có nhà hàng hiện có", "Function", "High", ["Mở modal sửa", "Đổi địa chỉ/phone/email", "Lưu"], "API trả cập nhật thành công"),
        tc("Cập nhật nhà hàng không tồn tại", pre, "Negative", "Medium", ["Gọi PUT /restaurants/999999"], "API trả 404 không tìm thấy nhà hàng"),
        tc("Duyệt nhà hàng đang chờ", "Có nhà hàng status Đang chờ duyệt", "Function", "High", ["Mở modal duyệt", "Nhấn Duyệt"], "Status chuyển Đang hoạt động"),
        tc("Từ chối nhà hàng đang chờ", "Có nhà hàng status Đang chờ duyệt", "Function", "High", ["Mở modal duyệt", "Nhấn Từ chối"], "Status chuyển Từ chối"),
        tc("Duyệt nhà hàng không tồn tại", pre, "Negative", "Medium", ["Gọi PUT /restaurants/999999/approve"], "API trả 404 không tìm thấy nhà hàng"),
        tc("Từ chối nhà hàng không tồn tại", pre, "Negative", "Medium", ["Gọi PUT /restaurants/999999/reject"], "API trả 404 không tìm thấy nhà hàng"),
        tc("Ẩn nhà hàng bằng thao tác xóa", "Có nhà hàng hiện có", "Function", "High", ["Nhấn Xóa", "Xác nhận"], "Restaurant không bị xóa vật lý, status chuyển Ngưng hoạt động"),
        tc("Xóa nhà hàng không tồn tại", pre, "Negative", "Medium", ["Gọi DELETE /restaurants/999999"], "API trả lỗi phù hợp"),
        tc("Chặn tạo nhà hàng khi không phải ADMIN", "Đăng nhập STAFF hoặc CUSTOMER", "Security", "Critical", ["Gọi POST /admin/restaurants"], "API trả 403"),
        tc("Chặn duyệt nhà hàng khi không phải ADMIN", "Đăng nhập STAFF", "Security", "Critical", ["Gọi PUT /restaurants/{id}/approve"], "API trả 403"),
        tc("Danh sách chủ nhà hàng tải vào combobox", pre, "Integration", "Medium", ["Mở modal thêm/sửa", "Quan sát select chủ nhà hàng"], "Select hiển thị user có thể gán làm chủ"),
        tc("Danh sách cuisine tải vào combobox", pre, "Integration", "Medium", ["Mở modal thêm/sửa", "Quan sát select cuisine"], "Select hiển thị CuisineID/CuisineName từ API"),
        tc("Đăng ký nhà hàng bởi STAFF chờ duyệt", "Đăng nhập STAFF", "Integration", "High", ["Gọi /registerRestaurant với dữ liệu hợp lệ", "Admin mở danh sách"], "Nhà hàng mới có status Đang chờ duyệt để admin duyệt"),
    ]


def admin_report_cases():
    pre = "Admin đã đăng nhập và có dữ liệu Payment/Reservation"
    return [
        tc("Tải báo cáo tất cả nhà hàng theo tháng", pre, "Function", "Critical", ["Mở report.html", "Chọn tháng", "Nhấn Xem báo cáo"], "API trả status success, total_report, total_6_months và danh sách nhà hàng"),
        tc("Tải báo cáo một nhà hàng cụ thể", pre, "Function", "High", ["Chọn một restaurant_id", "Chọn tháng", "Nhấn Xem báo cáo"], "Chỉ hiển thị doanh thu của nhà hàng đã chọn"),
        tc("Thiếu tháng báo cáo", pre, "Validation", "High", ["Xóa report_month", "Nhấn Xem báo cáo"], "Frontend/API báo thiếu tháng cần xem báo cáo"),
        tc("Chặn báo cáo khi không phải ADMIN", "Đăng nhập STAFF", "Security", "Critical", ["Gọi POST /admin/report"], "API trả 403"),
        tc("Chặn báo cáo khi thiếu token", "Không có token", "Security", "High", ["Gọi POST /admin/report"], "API trả 401"),
        tc("Báo cáo tháng không có doanh thu", "Không có Payment trong tháng chọn", "Boundary", "Medium", ["Chọn tháng không có payment", "Tải báo cáo"], "Tổng doanh thu bằng 0, bảng vẫn hiển thị nhà hàng phù hợp"),
        tc("Báo cáo tính doanh thu tháng chọn", pre, "Function", "High", ["Có Payment trong tháng", "Tải báo cáo"], "selected_month_revenue bằng tổng Payment.Amount của tháng"),
        tc("Báo cáo tính tổng 6 tháng", pre, "Function", "High", ["Có Payment trong 6 tháng gần nhất", "Tải báo cáo"], "total_6_months bằng tổng doanh thu 6 tháng"),
        tc("Báo cáo tạo đủ 6 cột tháng", pre, "UI", "Medium", ["Tải báo cáo", "Quan sát header bảng"], "Bảng có các cột tháng trong 6 tháng gần nhất"),
        tc("Sắp xếp nhà hàng theo doanh thu tháng chọn", pre, "Function", "Medium", ["Tải báo cáo nhiều nhà hàng", "Quan sát thứ tự"], "Nhà hàng doanh thu cao hơn đứng trước"),
        tc("Không tính nhà hàng Ngưng hoạt động", "Có nhà hàng status Ngưng hoạt động có payment", "Business", "High", ["Tải báo cáo"], "Nhà hàng ngưng hoạt động không nằm trong kết quả"),
        tc("Tải danh sách nhà hàng cho filter", pre, "Integration", "Medium", ["Mở report.html", "Gọi GET /admin/restaurants"], "Combobox có Tất cả nhà hàng và từng nhà hàng"),
        tc("Định dạng tiền VNĐ", pre, "UI", "Low", ["Tải báo cáo", "Quan sát các số tiền"], "Tiền hiển thị theo locale vi-VN và hậu tố VNĐ"),
        tc("Báo cáo restaurant_id không tồn tại", pre, "Negative", "Medium", ["Gửi restaurant_id=999999", "Chọn tháng hợp lệ"], "Kết quả rỗng, total = 0, không lỗi server"),
        tc("Báo cáo report_month sai định dạng", pre, "Negative", "High", ["Gửi report_month=2026/04", "Quan sát phản hồi"], "Hệ thống cần trả lỗi dữ liệu thay vì lỗi 500"),
        tc("Báo cáo với payment CreatedAt rỗng", "Có Payment thiếu CreatedAt", "Boundary", "Medium", ["Tải báo cáo"], "Payment thiếu CreatedAt không được tính"),
        tc("Báo cáo cập nhật sau thanh toán mới", "Có booking vừa thanh toán đặt cọc", "Integration", "High", ["Tạo Payment trong tháng", "Tải lại báo cáo"], "Doanh thu tăng đúng Amount"),
        tc("Hiển thị trạng thái loading khi tải báo cáo", pre, "UI", "Low", ["Nhấn Xem báo cáo", "Quan sát nút và text trạng thái"], "Nút bị disabled tạm thời và hiện trạng thái đang tải"),
        tc("Hiển thị lỗi khi token hết hạn", "Token admin hết hạn", "Security", "High", ["Tải báo cáo"], "Frontend hiển thị token không hợp lệ hoặc yêu cầu đăng nhập lại"),
        tc("Bảng rỗng khi không có nhà hàng phù hợp", "Không có nhà hàng hoạt động trong bộ lọc", "Boundary", "Low", ["Tải báo cáo"], "Bảng hiển thị thông báo không có dữ liệu doanh thu"),
    ]


def clone_sheet_format(wb, source_ws, name):
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)

    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))

    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    if ws.max_column:
        ws.delete_cols(1, ws.max_column)

    for col in range(1, 12):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = source_ws.column_dimensions[letter].width

    for row_idx in range(1, 6):
        ws.row_dimensions[row_idx].height = source_ws.row_dimensions[row_idx].height
        for col in range(1, 12):
            src = source_ws.cell(row_idx, col)
            dst = ws.cell(row_idx, col)
            if src.has_style:
                dst._style = copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            if src.alignment:
                dst.alignment = copy(src.alignment)

    return ws


def write_sheet(ws, source_ws, prefix, cases):
    ws.title = ws.title
    ws["A1"] = "Test Cases"
    ws.merge_cells("A1:K1")
    ws["A1"].font = copy(source_ws["A1"].font)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(5, col)
        cell.value = header
        cell.font = copy(source_ws.cell(5, min(col, source_ws.max_column)).font) if source_ws.max_column >= col else Font(bold=True)
        cell.font = Font(name=cell.font.name, size=cell.font.sz, bold=True, color=cell.font.color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if source_ws.cell(5, min(col, source_ws.max_column)).has_style:
            cell._style = copy(source_ws.cell(5, min(col, source_ws.max_column))._style)
            cell.font = Font(name=cell.font.name, size=cell.font.sz, bold=True, color=cell.font.color)

    row = 6
    body_style_row = 6
    for idx, case in enumerate(cases, start=1):
        start = row
        steps = case["steps"]
        for step_no, step in enumerate(steps, start=1):
            for col in range(1, 12):
                src = source_ws.cell(body_style_row, min(col, source_ws.max_column))
                dst = ws.cell(row, col)
                if src.has_style:
                    dst._style = copy(src._style)
                dst.alignment = Alignment(vertical="top", wrap_text=True)
            if step_no == 1:
                ws.cell(row, 1).value = f"{prefix}_{idx}"
                ws.cell(row, 2).value = case["desc"]
                ws.cell(row, 3).value = case["prereq"]
                ws.cell(row, 4).value = case["type"]
                ws.cell(row, 5).value = case["level"]
                ws.cell(row, 8).value = case["expected"]
            ws.cell(row, 6).value = step_no
            ws.cell(row, 7).value = step
            row += 1

        end = row - 1
        if end > start:
            for col in [1, 2, 3, 4, 5, 8, 9, 10, 11]:
                ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)

    max_row = row - 1
    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = 28 if r >= 6 else ws.row_dimensions[r].height

    for col in range(1, 12):
        letter = get_column_letter(col)
        max_len = len(str(ws.cell(5, col).value or ""))
        for r in range(6, max_row + 1):
            value = ws.cell(r, col).value
            if value:
                max_len = max(max_len, max(len(line) for line in str(value).splitlines()))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 55)

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:K{max_row}"


def workbook_summary(wb):
    return {
        ws.title: {
            "rows": ws.max_row,
            "cols": ws.max_column,
            "test_cases": sum(
                1 for row in range(6, ws.max_row + 1)
                if isinstance(ws.cell(row, 1).value, str) and re.search(r"_\d+$", ws.cell(row, 1).value)
            ),
            "headers": [ws.cell(5, c).value for c in range(1, 12)],
        }
        for ws in wb.worksheets
    }


def main():
    add_sheet_cases()
    wb = load_workbook(TARGET)
    source_ws = wb["LOGIN_REGISTER"] if "LOGIN_REGISTER" in wb.sheetnames else wb.worksheets[0]

    for sheet_name, (prefix, cases) in SHEET_CASES.items():
        ws = clone_sheet_format(wb, source_ws, sheet_name)
        write_sheet(ws, source_ws, prefix, cases[:20])

    wb.save(TARGET)
    print(workbook_summary(wb))


if __name__ == "__main__":
    main()
