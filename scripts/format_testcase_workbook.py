import re
import sys
from copy import copy
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / ".codex_deps"))

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange


TARGET = Path(r"C:\Users\nhatl\OneDrive\Tài liệu\Downloads\TestCase_RestaurantBookingSystem.xlsx")
TABLE_COLS = 11
HEADER_ROW = 5
DATA_START = 6
MERGE_COLS = [1, 2, 3, 4, 5, 8, 9, 10, 11]


def is_test_case_id(value):
    return isinstance(value, str) and bool(re.search(r"_\d+$", value.strip()))


def find_last_table_row(ws):
    last = HEADER_ROW
    for row in range(DATA_START, ws.max_row + 1):
        if any(ws.cell(row, col).value not in (None, "") for col in range(1, TABLE_COLS + 1)):
            last = row
    return last


def remove_blank_table_rows(ws):
    for row in range(ws.max_row, DATA_START - 1, -1):
        if not any(ws.cell(row, col).value not in (None, "") for col in range(1, TABLE_COLS + 1)):
            ws.delete_rows(row, 1)


def test_case_ranges(ws, last_row):
    starts = [
        row
        for row in range(DATA_START, last_row + 1)
        if is_test_case_id(ws.cell(row, 1).value)
    ]
    ranges = []
    for i, start in enumerate(starts):
        end = (starts[i + 1] - 1) if i + 1 < len(starts) else last_row
        while end > start and not any(
            ws.cell(end, col).value not in (None, "") for col in range(1, TABLE_COLS + 1)
        ):
            end -= 1
        ranges.append((start, end))
    return ranges


def unmerge_table_area(ws, last_row):
    for merged in list(ws.merged_cells.ranges):
        cr = CellRange(str(merged))
        if cr.min_row >= DATA_START and cr.max_row <= max(last_row, DATA_START) and cr.min_col <= TABLE_COLS:
            ws.unmerge_cells(str(merged))


def apply_header_style(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    for col in range(1, TABLE_COLS + 1):
        cell = ws.cell(HEADER_ROW, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def apply_table_style(ws, last_row):
    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in range(HEADER_ROW, last_row + 1):
        for col in range(1, TABLE_COLS + 1):
            cell = ws.cell(row, col)
            cell.border = border
            horizontal = "center" if col in [1, 4, 5, 6, 10] else "left"
            if row == HEADER_ROW:
                horizontal = "center"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center" if row == HEADER_ROW else "top", wrap_text=True)


def merge_case_rows(ws, ranges):
    for start, end in ranges:
        if end <= start:
            continue
        for col in MERGE_COLS:
            ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)


def autofit(ws, last_row):
    for col in range(1, TABLE_COLS + 1):
        max_len = 0
        for row in range(1, last_row + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue
            for line in str(value).splitlines():
                max_len = max(max_len, len(line))
        width = min(max(max_len + 2, 10), 55)
        if col == 6:
            width = 10
        if col in [1, 4, 5, 10]:
            width = min(max(width, 12), 20)
        if col in [2, 3, 7, 8]:
            width = min(max(width, 25), 55)
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[HEADER_ROW].height = 28
    for row in range(DATA_START, last_row + 1):
        ws.row_dimensions[row].height = 30


def format_sheet(ws):
    if ws.max_row < HEADER_ROW or ws.max_column < TABLE_COLS:
        return 0
    headers = [ws.cell(HEADER_ROW, col).value for col in range(1, TABLE_COLS + 1)]
    if headers[0] != "Test Case ID":
        return 0

    unmerge_table_area(ws, ws.max_row)
    remove_blank_table_rows(ws)
    last_row = find_last_table_row(ws)
    ranges = test_case_ranges(ws, last_row)
    merge_case_rows(ws, ranges)
    apply_header_style(ws)
    apply_table_style(ws, last_row)
    autofit(ws, last_row)
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{HEADER_ROW}:K{last_row}"
    return len(ranges)


def main():
    wb = load_workbook(TARGET)
    counts = {}
    for ws in wb.worksheets:
        counts[ws.title] = format_sheet(ws)
    wb.save(TARGET)
    print(counts)


if __name__ == "__main__":
    main()
